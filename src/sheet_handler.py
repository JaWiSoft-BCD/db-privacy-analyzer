# src/output/excel_generator.py

import pandas as pd
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
from utils import append_to_file
import logging
from dataclasses import dataclass
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import os

@dataclass
class ExcelStyling:
    """Defines the styling configuration for Excel sheets"""
    header_fill: str = "1F4E79"
    header_font_color: str = "FFFFFF"
    alternate_row_fill: str = "F2F2F2"
    border_color: str = "CCCCCC"
    highlight_color: str = "FFE699"

class ExcelGenerator:
    def __init__(self, output_dir: str = "reports"):
        """Initialize Excel generator with output directory."""
        self.output_dir = Path(output_dir)
        self.ensure_directories()
        self.styling = ExcelStyling()
        self._setup_logging()
        self._ensure_output_dir()

    def ensure_directories(self):
        """Create input and output directories if they don't exist."""
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _setup_logging(self):
        """Configure logging for Excel operations."""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)

    def _ensure_output_dir(self):
        """Ensure output directory exists."""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)

    def generate_report(self, ai_analysis: List[Dict[str, Any]], database_name: str) -> str:
        """
        Generate an Excel report from AI analysis results.
        Returns the path to the generated file.
        """
        try:
            # Create timestamp for filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"{database_name}_privacy_analysis_{timestamp}.xlsx"
            filepath = os.path.join(self.output_dir, filename)

            # Create Excel writer object
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # self._create_summary_sheet(ai_analysis, writer)
                self._create_detailed_analysis_sheet(ai_analysis, writer)
                # self._create_privacy_impact_sheet(ai_analysis, writer)
                # self._create_recommendations_sheet(ai_analysis, writer)

                # Apply styling to all sheets
                for sheet_name in writer.sheets:
                    self._apply_sheet_styling(writer.sheets[sheet_name])

            self.logger.info(f"Successfully generated report: {filepath}")
            return filepath

        except Exception as e:
            self.logger.error(f"Error generating Excel report: {str(e)}")
            raise

    def _create_summary_sheet(self, ai_analysis: List[Dict[str, Any]], writer: pd.ExcelWriter):
        """Create summary sheet with high-level overview."""
        summary_data = {
            'Category': [],
            'Count': [],
            'Risk Level': [],
            'Description': []
        }

        # Process AI analysis to create summary
        data_categories = set()
        risk_levels = {'High': 0, 'Medium': 0, 'Low': 0}
        
        for item in ai_analysis:
            data_categories.add(item.get('data_category', 'Uncategorized'))
            risk_levels[item.get('risk_level', 'Low')] += 1

        # Add summary rows
        summary_data['Category'].extend(['Data Categories', 'High Risk Items', 'Medium Risk Items', 'Low Risk Items'])
        summary_data['Count'].extend([
            len(data_categories),
            risk_levels['High'],
            risk_levels['Medium'],
            risk_levels['Low']
        ])
        summary_data['Risk Level'].extend(['N/A', 'High', 'Medium', 'Low'])
        summary_data['Description'].extend([
            'Total unique data categories found',
            'Items requiring immediate attention',
            'Items requiring regular review',
            'Items with minimal privacy impact'
        ])

        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)

    def _create_detailed_analysis_sheet(self, ai_analysis: List[Dict[str, Any]], writer: pd.ExcelWriter):
        """
        Create a detailed analysis sheet with all findings.

        Args:
            ai_analysis (List[Dict[str, Any]]): The AI analysis output, where each dict represents the analysis for a table.
            writer (pd.ExcelWriter): The ExcelWriter instance to write the sheet to.

        Raises:
            ValueError: If a required field is missing from the AI analysis output.
        """
        detailed_data = []
        # Define the required fields and their default values
        required_fields = {
            'Column': '',
            'Description': '',
            'Data Type': '',
            'Collection Method': '',
            'Data Source': '',
            'Primary Purpose': '',
            'Legal Basis': '',
            'Personal Data': '',
            'Personal Information': ''
        }

        for item in ai_analysis:
            table = item.get('table_name', '')
            for column_item in item["column_report"]:
                row_data = {}
                row_data['Table'] = table
                for field, default_value in required_fields.items():
                    value = column_item.get(field, default_value)
                    if value is None:
                        append_to_file("Ai Analysis errors.txt", f"Required field '{field}' for table: {table} is missing from the AI analysis output.")
                        print(f"Required field '{field}' for table: {table} is missing from the AI analysis output.")
                    row_data[field] = value
                detailed_data.append(row_data)

        df_detailed = pd.DataFrame(detailed_data)
        df_detailed.to_excel(writer, sheet_name='Detailed Analysis', index=False)

    def _create_privacy_impact_sheet(self, ai_analysis: List[Dict[str, Any]], writer: pd.ExcelWriter):
        """Create privacy impact assessment sheet."""
        impact_data = {
            'Impact Area': [],
            'Description': [],
            'Affected Data': [],
            'Mitigation Measures': []
        }

        # Group findings by privacy impact
        impact_areas = {}
        for item in ai_analysis:
            impact = item.get('privacy_impact', '')
            if impact not in impact_areas:
                impact_areas[impact] = {
                    'affected_data': set(),
                    'mitigation': set()
                }
            impact_areas[impact]['affected_data'].add(f"{item.get('table_name', '')}.{item.get('column_name', '')}")
            if 'mitigation_steps' in item:
                impact_areas[impact]['mitigation'].update(item['mitigation_steps'])

        # Create rows for each impact area
        for impact, details in impact_areas.items():
            impact_data['Impact Area'].append(impact)
            impact_data['Description'].append(self._generate_impact_description(impact))
            impact_data['Affected Data'].append('\n'.join(details['affected_data']))
            impact_data['Mitigation Measures'].append('\n'.join(details['mitigation']))

        df_impact = pd.DataFrame(impact_data)
        df_impact.to_excel(writer, sheet_name='Privacy Impact', index=False)

    def _create_recommendations_sheet(self, ai_analysis: List[Dict[str, Any]], writer: pd.ExcelWriter):
        """Create recommendations sheet with actionable items."""
        recommendations_data = {
            'Priority': [],
            'Recommendation': [],
            'Affected Items': [],
            'Implementation Steps': [],
            'Expected Outcome': []
        }

        # Extract and organize recommendations
        recommendations = self._extract_recommendations(ai_analysis)
        
        for rec in recommendations:
            recommendations_data['Priority'].append(rec['priority'])
            recommendations_data['Recommendation'].append(rec['recommendation'])
            recommendations_data['Affected Items'].append(rec['affected_items'])
            recommendations_data['Implementation Steps'].append(rec['implementation_steps'])
            recommendations_data['Expected Outcome'].append(rec['expected_outcome'])

        df_recommendations = pd.DataFrame(recommendations_data)
        df_recommendations.to_excel(writer, sheet_name='Recommendations', index=False)

    def _apply_sheet_styling(self, worksheet: Worksheet):
        """Apply consistent styling to worksheet."""
        # Define styles
        header_fill = PatternFill(start_color=self.styling.header_fill, end_color=self.styling.header_fill, fill_type="solid")
        header_font = Font(color=self.styling.header_font_color, bold=True)
        alternate_fill = PatternFill(start_color=self.styling.alternate_row_fill, end_color=self.styling.alternate_row_fill, fill_type="solid")
        border = Border(
            left=Side(style='thin', color=self.styling.border_color),
            right=Side(style='thin', color=self.styling.border_color),
            top=Side(style='thin', color=self.styling.border_color),
            bottom=Side(style='thin', color=self.styling.border_color)
        )

        # Apply header styling
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Apply alternate row coloring and borders
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if row_idx % 2 == 0:
                    cell.fill = alternate_fill

        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _generate_impact_description(self, impact: str) -> str:
        """Generate a description for a privacy impact."""
        impact_descriptions = {
            'High': 'Significant privacy concerns requiring immediate attention',
            'Medium': 'Moderate privacy concerns requiring regular monitoring',
            'Low': 'Minor privacy concerns with limited potential impact'
        }
        return impact_descriptions.get(impact, 'Impact level not specified')

    def _extract_recommendations(self, ai_analysis: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Extract and prioritize recommendations from AI analysis."""
        recommendations = []
        seen_recommendations = set()

        for item in ai_analysis:
            if 'recommended_action' in item:
                action = item['recommended_action']
                if action not in seen_recommendations:
                    recommendation = {
                        'priority': item.get('risk_level', 'Low'),
                        'recommendation': action,
                        'affected_items': f"{item.get('table_name', '')}.{item.get('column_name', '')}",
                        'implementation_steps': item.get('implementation_steps', 'Steps not provided'),
                        'expected_outcome': item.get('expected_outcome', 'Outcome not specified')
                    }
                    recommendations.append(recommendation)
                    seen_recommendations.add(action)
                else:
                    # Update affected items for existing recommendation
                    for rec in recommendations:
                        if rec['recommendation'] == action:
                            rec['affected_items'] += f"\n{item.get('table_name', '')}.{item.get('column_name', '')}"

        return sorted(recommendations, key=lambda x: {'High': 0, 'Medium': 1, 'Low': 2}[x['priority']])

# Usage example:
if __name__ == "__main__":
    # Sample AI analysis output
    sample_analysis = [
        {
            "table_name": "users",
            "column_name": "email",
            "data_category": "Personal Identifiable Information",
            "risk_level": "High",
            "privacy_impact": "Direct user identification",
            "recommended_action": "Implement encryption",
            "mitigation_steps": ["Enable column-level encryption", "Add access controls"],
            "implementation_steps": "1. Configure encryption\n2. Update access policies",
            "expected_outcome": "Enhanced data protection"
        }
    ]

    # Generate report
    excel_gen = ExcelGenerator()
    report_path = excel_gen.generate_report(sample_analysis, "sample_database")
    print(f"Report generated: {report_path}")